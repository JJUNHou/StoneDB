import requests
import subprocess
import shutil
import chromedriver_autoinstaller
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver  # Import from seleniumwire
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl
import json
import numpy as np
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import random
import os
from haversine import haversine
from selenium.webdriver.common.keys import Keys


## 엑셀 파일 오픈
workingdirectory = 'C:\\Users\\82103\\Desktop\\python\\tiger'
savefilename = workingdirectory + "\\tiger.xlsx"
logfilename = savefilename + "_log.txt"

#디버거크롬 가동
subprocess.Popen(r'C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\\chrometemp"') # 디버거 크롬 구동
options = webdriver.ChromeOptions()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
Chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
try:
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
except:
    chromedriver_autoinstaller.install(True)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.implicitly_wait(3)

rokdict= dict(zip(['basalt', 'granite','marble', 'limestone', 'sandstone'], ['3','4','5','6','7']))
r_rokdict = {v:k for k,v in rokdict.items()}
rokkey = list(rokdict.keys())
rokvalue = list(rokdict.values())

# 통합 문서 객체 생성
wb = openpyxl.Workbook()
# 시트 선택
ws = wb.active
ws.cell(row=1,column=1,value="링크")

url = f'https://www.tigerstone.co.kr/pro/main_list.php?gubun=1&aci_code=A8&'\
    f'cate_name=%C7%F6%B9%AB%BE%CF'

rowCount = 2
driver.get(url)
for i in range(3,8):
    try:
        
        roktype = driver.find_element(By.XPATH,'/html/body/table/tbody/tr/td/table/tbody/tr[4]/td/table/tbody/tr/td[1]/table/tbody/tr['+str(i)+']/td/a/img')
        roktype.click()
        sleep(1)

        for j in range(1,7):
            color = driver.find_element(By.XPATH,'/html/body/table/tbody/tr/td/table/tbody/tr[4]/td/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr/td['+str(2*j-1)+']')                                                    
            color.click()
            sleep(1)

            for k in range(1,10):
                try:
                    for n in range(1,3):
                        for m in range(1,7):
                            #sleep(0.3)
                            num = ws.cell(row=rowCount,column=1)
                            rok = ws.cell(row=rowCount,column=2)
                            name = ws.cell(row=rowCount,column=3)
                            origin = ws.cell(row=rowCount,column=4)
                            color = ws.cell(row=rowCount,column=5)
                            appli = ws.cell(row=rowCount,column=6)
                            img = ws.cell(row=rowCount,column=7)

                            name_tag = driver.find_element(By.XPATH,"/html/body/table/tbody/tr/td/table/tbody/tr[4]/td/table/tbody/tr/td[2]/table/tbody/tr[3]/td/table/tbody/tr[" + str(m)+"]/td["+ str(2*n-1) +"]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr[2]/td/strong/a").text
                            origin_tag = driver.find_element(By.XPATH,"/html/body/table/tbody/tr/td/table/tbody/tr[4]/td/table/tbody/tr/td[2]/table/tbody/tr[3]/td/table/tbody/tr["+ str(m)+"]/td["+str(2*n-1)+"]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr[3]/td").text
                            color_tag = driver.find_element(By.XPATH,"/html/body/table/tbody/tr/td/table/tbody/tr[4]/td/table/tbody/tr/td[2]/table/tbody/tr[3]/td/table/tbody/tr["+ str(m)+"]/td["+str(2*n-1)+"]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr[4]/td").text
                            appli_tag = driver.find_element(By.XPATH,"/html/body/table/tbody/tr/td/table/tbody/tr[4]/td/table/tbody/tr/td[2]/table/tbody/tr[3]/td/table/tbody/tr["+ str(m)+"]/td["+str(2*n-1)+"]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr[5]/td").text
                            img_tag = driver.find_element(By.XPATH,"/html/body/table/tbody/tr/td/table/tbody/tr[4]/td/table/tbody/tr/td[2]/table/tbody/tr[3]/td/table/tbody/tr["+str(m)+"]/td["+str(2*n-1)+"]/table/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr/td/a/img")
                            imgurl = img_tag.get_attribute('src')

                            num.value = rowCount-1
                            rok.value = r_rokdict[str(i)]
                            name.value = name_tag
                            origin.value = origin_tag.split(":")[1]
                            color.value = color_tag.split(":")[1]
                            appli.value = appli_tag.split(":")[1]
                            img.value = f"{imgurl}"

                            print(f"{rowCount}|{rok.value}|{name.value}|{origin.value}|{color.value}|{appli.value}{img.value}")
                            rowCount += 1

                    next = driver.find_element(By.XPATH,"/html/body/table/tbody/tr/td/table/tbody/tr[4]/td/table/tbody/tr/td[2]/table/tbody/tr[6]/td/table/tbody/tr[2]/td/font["+str(k)+"]/a")
                    next.click()
                except:
                    print("오류발생")

    except:
        print(f"오류발생")

wb.save(savefilename)
